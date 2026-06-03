"""
Build the RTLDI ATLAS master table for UN Member States.

Joins:
- UN list (for canonical 193)
- World Bank latest (G0 = gdp_pc, population, socio proxies for indicator #9)
- (Future) V-Dem subset for the other 8 components -> real R

For the initial version (pre full V-Dem integration):
- Computes a *placeholder* R using:
  - Socio component (real from WB undernourish + poverty) for #9
  - Average of 0.70 for the other 8 (near global source avg 0.62, slightly optimistic)
  - Or a simple poverty-modulated R for demo variation.
- Always uses the exact equation from the source.
- Outputs ranked table with per-capita and total deficit.
- Also writes a nicely formatted XLSX (openpyxl).

Run:
  python -m src.build_atlas --year 2024 --eta 0.05 --output outputs/atlas/rtl_di_atlas_2024.csv
"""

from __future__ import annotations
import argparse
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from .rtl_di import (
    compute_delta_g,
    total_deficit,
    compute_rtlp_from_binaries,
    score_socioeconomic,
    RTLDIConfig,
)
from .fetch_vdem import get_vdem_r_for_iso3s, load_vdem_for_rtlp, compute_rtlp_vdem_row


def load_un_members(path: str = "data/raw/un_member_states.csv") -> pd.DataFrame:
    df = pd.read_csv(path)
    # Ensure consistent columns
    df = df.rename(columns={"name_short": "country", "ISO3": "iso3"})
    return df[["iso3", "country", "UNregion"]].copy()


def load_wb_latest(path: str = "data/processed/wb_un_members_2023_2024_latest.csv") -> pd.DataFrame:
    df = pd.read_csv(path)
    # Standardize
    df = df.rename(columns={"economy": "iso3", "Country": "wb_name"})
    # Keep useful
    keep = ["iso3", "gdp_pc_current", "population", "undernourish_pct", "poverty_215_pct", "atlas_year"]
    for c in keep:
        if c not in df.columns:
            df[c] = np.nan
    return df[keep].copy()


def load_local_wb_gdp_for_year(year: int = 2023, base_dir: str = "data/raw/wb_gdp") -> pd.DataFrame:
    """Load G0 (GDP pc current US$) from the user's downloaded WB API bulk CSV for a specific year."""
    csv = f"{base_dir}/API_NY.GDP.PCAP.CD_DS2_en_csv_v2_46.csv"
    df = pd.read_csv(csv, skiprows=4)
    col = str(year)
    if col not in df.columns:
        col = str(int(year))
    gdp = df[["Country Code", col]].rename(columns={"Country Code": "iso3", col: "g0"})
    gdp["iso3"] = gdp["iso3"].astype(str).str.upper()
    gdp = gdp.dropna(subset=["g0"])
    gdp["g0_year"] = year
    return gdp[["iso3", "g0", "g0_year"]]


def compute_real_r(
    iso3: str,
    vdem_row: Optional[pd.Series],
    undernourish_pct: Optional[float],
    poverty_pct: Optional[float],
    cfg: Optional[RTLDIConfig] = None,
) -> float:
    """Full 0-1 R from 8 V-Dem components (if row provided) + real socio #9."""
    cfg = cfg or RTLDIConfig()
    if vdem_row is not None and len(vdem_row) > 0:
        r8 = compute_rtlp_vdem_row(vdem_row, cfg)  # 0-1 for the 8
        r_from_vdem = r8 * (8 / 9.0)
    else:
        r_from_vdem = 0.0
    socio_bin = score_socioeconomic(undernourish_pct, poverty_pct, cfg)
    r_from_socio = socio_bin / 9.0
    r = r_from_vdem + r_from_socio
    return float(max(0.0, min(1.0, r)))


def compute_placeholder_r(
    row: pd.Series,
    base_r_other: float = 0.70,
    config: Optional[RTLDIConfig] = None,
) -> float:
    """Legacy placeholder (for comparison). Prefer compute_real_r with V-Dem."""
    cfg = config or RTLDIConfig()
    socio_bin = score_socioeconomic(row.get("undernourish_pct"), row.get("poverty_215_pct"), cfg)
    r_from_socio = socio_bin / 9.0
    r_from_others = base_r_other * (8 / 9.0)
    r = r_from_others + r_from_socio
    poverty = row.get("poverty_215_pct") or 0.0
    under = row.get("undernourish_pct") or 0.0
    penalty = min(0.15, (poverty / 100.0) * 0.3 + (under / 100.0) * 0.2)
    r = max(0.0, min(1.0, r - penalty))
    return float(r)


def build_atlas(
    year: int = 2026,
    eta: float = 0.05,
    un_path: str = "data/raw/un_member_states.csv",
    wb_path: str = "data/processed/wb_un_members_2023_2024_latest.csv",
    out_dir: str = "outputs/atlas",
    use_real_vdem: bool = True,
) -> pd.DataFrame:
    """Build RTLDI ATLAS. When use_real_vdem=True (default), uses the 9-component R from V-Dem + WB socio."""
    un = load_un_members(un_path)

    vdem_year_used = 2024 if year >= 2025 else year

    # Socio + pop from previous processed (or could call wbgapi here)
    wb = load_wb_latest(wb_path)

    # Join UN + socio/pop
    df = un.merge(wb, on="iso3", how="left")

    # For 2026+ atlas: MUST use the freshest available GDP (most dynamic variable),
    # even if V-Dem RTLP components are from 2024. Load the pre-fetched latest GDP table.
    vdem_year_used = 2024
    if year >= 2025:
        try:
            g0_latest = pd.read_csv("data/processed/wb_gdp_latest_for_2026_atlas.csv")
            print(f"Using fresh latest GDP data for {year} atlas baseline (from WB API mrv): {len(g0_latest)} countries")
            g0m = g0_latest[["iso3", "g0_latest", "g0_year"]].rename(columns={"g0_latest": "g0"})
            df = df.merge(g0m, on="iso3", how="left")
            df["g0"] = df["g0"].fillna(df.get("gdp_pc_current"))
            df["g0_year"] = df["g0_year"].fillna(2026)
        except Exception as e:
            print(f"Could not load fresh 2026 GDP table ({e}), falling back.")
            df["g0"] = df.get("gdp_pc_current")
            df["g0_year"] = 2026
    else:
        # For historical years, prefer local WB GDP the user provided for G0
        try:
            g0_df = load_local_wb_gdp_for_year(year)
            print(f"Using local WB GDP CSV for G0 ({year}): {len(g0_df)} countries")
            if len(g0_df):
                g0m = g0_df[["iso3", "g0"]].rename(columns={"g0": "g0_local"})
                df = df.merge(g0m, on="iso3", how="left")
                df["g0"] = df["g0_local"].fillna(df.get("gdp_pc_current"))
                df = df.drop(columns=["g0_local"], errors="ignore")
            else:
                df["g0"] = df.get("gdp_pc_current")
        except Exception:
            df["g0"] = df.get("gdp_pc_current")
            df["g0_year"] = year
            print("Local WB GDP not found for historical year.")

    if use_real_vdem:
        print(f"Loading real V-Dem components for year {year} ...")
        vdem_r = get_vdem_r_for_iso3s(year=year, iso3s=un["iso3"].tolist())
        print("V-Dem R loaded for", len(vdem_r), "countries")
        # Merge vdem r8
        df = df.merge(vdem_r[["iso3", "r_vdem8"]], on="iso3", how="left")
        # For each row, get the vdem raw row if possible for full compute (or use precomputed r8 + socio)
        # Simpler: load full components df once
        vdem_full = load_vdem_for_rtlp(year)
        vdem_full["iso3"] = vdem_full["country_text_id"].astype(str).str.upper()
        vdem_full = vdem_full.set_index("iso3")

        def _real_r_for_row(row):
            iso = row["iso3"]
            vrow = vdem_full.loc[iso] if iso in vdem_full.index else None
            return compute_real_r(
                iso,
                vrow,
                row.get("undernourish_pct"),
                row.get("poverty_215_pct"),
            )

        df["r"] = df.apply(_real_r_for_row, axis=1)
        r_col = "r"
        note = f"R from 8 V-Dem components (crosswalk, year={vdem_year_used}) + real WB socio #9 + 2026-fresh G0 (GDP is the more dynamic variable). See docs/indicator_crosswalk.md"
        df["vdem_year"] = vdem_year_used
    else:
        df["r_placeholder"] = df.apply(compute_placeholder_r, axis=1)
        r_col = "r_placeholder"
        note = "PLACEHOLDER R (socio real + avg 0.70 on other 8)."
        df["vdem_year"] = vdem_year_used
    if "g0_year" not in df.columns:
        df["g0_year"] = year

    # ΔG per capita using the chosen r
    df["delta_g_per_capita"] = df.apply(
        lambda row: compute_delta_g(row[r_col], row["g0"] if pd.notna(row["g0"]) else 0.0, eta)
        if pd.notna(row["g0"]) and pd.notna(row[r_col]) else np.nan,
        axis=1,
    )

    # Total
    df["total_deficit_usd"] = df.apply(
        lambda row: total_deficit(row["delta_g_per_capita"], row["population"])
        if pd.notna(row["delta_g_per_capita"]) and pd.notna(row["population"])
        else np.nan,
        axis=1,
    )

    # Ranks
    df["rank_by_total_deficit"] = df["total_deficit_usd"].rank(ascending=False, method="min").astype("Int64")
    df["rank_by_r_lowest"] = df[r_col].rank(ascending=True, method="min").astype("Int64")

    # Flags + notes
    df["has_gdp"] = df["g0"].notna()
    df["has_pop"] = df["population"].notna()
    df["has_socio"] = df["undernourish_pct"].notna() | df["poverty_215_pct"].notna()
    df["data_notes"] = note

    # Master columns (use "r" unified)
    if "r" not in df.columns:
        df["r"] = df.get(r_col)
    master_cols = [
        "iso3", "country", "UNregion", "r", "g0", "g0_year", "vdem_year", "delta_g_per_capita", "population",
        "total_deficit_usd", "rank_by_total_deficit", "rank_by_r_lowest",
        "undernourish_pct", "poverty_215_pct", "has_gdp", "has_socio", "data_notes"
    ]
    master = df[[c for c in master_cols if c in df.columns]].copy()
    master = master.sort_values("rank_by_total_deficit")

    # Save
    out_dir_p = Path(out_dir)
    out_dir_p.mkdir(parents=True, exist_ok=True)
    suffix = "" if use_real_vdem else "_placeholder"
    csv_path = out_dir_p / f"rtl_di_atlas_un_members_{year}{suffix}.csv"
    master.to_csv(csv_path, index=False)
    print(f"Saved CSV: {csv_path}")

    summary = {
        "atlas_year": year,
        "eta": eta,
        "n_countries": len(master),
        "n_with_gdp": int(master["has_gdp"].sum()),
        "global_avg_r": float(master["r"].mean()) if "r" in master else float(master.get(r_col, pd.Series([np.nan])).mean()),
        "total_global_deficit_est_usd": float(master["total_deficit_usd"].sum()),
        "median_r": float(master["r"].median()) if "r" in master else float("nan"),
        "use_real_vdem": use_real_vdem,
        "note": note,
    }
    pd.Series(summary).to_json(out_dir_p / f"rtl_di_atlas_summary_{year}{suffix}.json", indent=2)
    print("Summary:", summary)

    xlsx_path = out_dir_p / f"rtl_di_atlas_un_members_{year}{suffix}.xlsx"
    write_xlsx_atlas(master, xlsx_path, year=year, eta=eta, summary=summary)
    print(f"Saved XLSX: {xlsx_path}")

    return master


def write_xlsx_atlas(
    df: pd.DataFrame,
    path: Path,
    year: int,
    eta: float,
    summary: dict,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"RTLDI ATLAS {year}"

    # Title
    ws.merge_cells("A1:O1")
    is_placeholder = "placeholder" in str(summary.get("note", "")).lower() or not summary.get("use_real_vdem", True)
    title_suffix = "(placeholder R)" if is_placeholder else "(real 9-component R from V-Dem + WB)"
    ws["A1"] = f"RTLDI ATLAS of UN Member States — {year} {title_suffix}"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:O2")
    note_text = summary.get("note", "R from V-Dem crosswalk + WB socio #9 per source document")
    ws["A2"] = (
        f"Source equation: ΔG = η × (1 − R) × G₀   |   η = {eta} (Hubbard, Causality and Attraction v3, DOI 10.5281/zenodo.19468550)   |   "
        f"R = {note_text}"
    )
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(wrap_text=True)

    # Summary row
    ws.merge_cells("A4:E4")
    avg_r_key = "global_avg_r" if "global_avg_r" in summary else "global_avg_r_placeholder"
    avg_r = summary.get(avg_r_key, summary.get("global_avg_r", float("nan")))
    ws["A4"] = f"Countries: {summary['n_countries']} | With GDP: {summary['n_with_gdp']} | Est. global deficit: ${summary['total_global_deficit_est_usd']:,.0f} | Avg R: {avg_r:.3f}"
    ws["A4"].font = Font(bold=True)

    # Headers
    headers = [
        "Rank (total $)", "ISO3", "Country", "UN Region", "R", "G0 (GDP pc current US$)",
        "ΔG per capita (US$)", "Population", "Total annual deficit (US$)", "Rank (lowest R)",
        "Undernourish %", "Poverty $2.15 %", "Has GDP", "Has Socio", "Notes",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for r_idx, row in enumerate(df.itertuples(index=False), start=7):
        def safe(v):
            if pd.isna(v):
                return None
            if isinstance(v, (pd.Int64Dtype, int)) and pd.isna(v):
                return None
            return v
        vals = [
            safe(row[8]),  # rank total
            row[0],  # iso3
            row[1],  # country
            row[2],  # region
            round(float(row[3]), 3) if pd.notna(row[3]) else None,  # r
            safe(row[4]),  # g0
            safe(row[5]),  # delta_g
            safe(row[6]),  # pop
            safe(row[7]),  # total
            safe(row[9]),  # rank r
            safe(row[10]), # under
            safe(row[11]), # pov
            "Y" if bool(row[12]) else "N",
            "Y" if bool(row[13]) else "N",
            str(row[14])[:60] + "..." if len(str(row[14])) > 60 else str(row[14]),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if c_idx in (1,5,6,9,10) else "left")
            # Number formats
            if c_idx in (6, 7, 9):  # money
                cell.number_format = '#,##0.00'
            if c_idx == 8:  # pop
                cell.number_format = '#,##0'
            if c_idx in (11,12):  # pct
                cell.number_format = '0.0'

    # Column widths
    widths = [12, 8, 28, 20, 14, 18, 18, 14, 20, 14, 12, 12, 8, 8, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Footer
    last = 7 + len(df)
    ws.merge_cells(f"A{last+2}:O{last+2}")
    ws[f"A{last+2}"] = "PLACEHOLDER: Full R from 9 RTLP indicators requires V-Dem v16 data (see docs/indicator_crosswalk.md and src/fetch_vdem.py). This demonstrates the equation + pipeline with real WB G0/pop/socio."
    ws[f"A{last+2}"].font = Font(italic=True, size=9, color="666666")

    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--eta", type=float, default=0.05)
    parser.add_argument("--output-dir", default="outputs/atlas")
    args = parser.parse_args()

    build_atlas(year=args.year, eta=args.eta, out_dir=args.output_dir)


if __name__ == "__main__":
    main()
